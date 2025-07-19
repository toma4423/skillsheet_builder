import io
from datetime import datetime
from typing import List, Optional

import openpyxl
from fastapi import FastAPI, File, UploadFile, HTTPException, Request
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel

app = FastAPI(title="SkillSheet Builder", description="ITエンジニア向けスキルシート作成Webアプリケーション")

# 静的ファイルとテンプレートの設定
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")


# データモデル
class BasicInfo(BaseModel):
    name: str
    kana: str
    gender: str
    age: Optional[int] = None
    nearest_station: Optional[str] = None
    experience_years: Optional[int] = None
    self_pr: Optional[str] = None
    main_technologies: Optional[str] = None
    qualifications: Optional[str] = None


class PossibleTasks(BaseModel):
    customer_negotiation: str = "-"
    research_analysis: str = "-"
    requirement_definition: str = "-"
    basic_design: str = "-"
    detailed_design: str = "-"
    pg_development: str = "-"
    unit_test: str = "-"
    integration_test: str = "-"
    system_maintenance: str = "-"
    nw_design: str = "-"
    nw_construction: str = "-"
    nw_operation: str = "-"
    sv_design: str = "-"
    sv_construction: str = "-"
    sv_operation: str = "-"


class CareerHistoryEntry(BaseModel):
    start_date: str
    end_date: str
    duration: Optional[str] = None
    overview: str
    position: str
    scale_members: Optional[int] = None
    responsibilities: str
    tech_environment: Optional[str] = None


class SkillSheetData(BaseModel):
    basic_info: BasicInfo
    possible_tasks: PossibleTasks
    career_history: List[CareerHistoryEntry]


# ルート
@app.get("/", response_class=HTMLResponse)
async def root(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/skillsheet", response_class=HTMLResponse)
async def skillsheet(request: Request):
    return templates.TemplateResponse("skillsheet.html", {"request": request})


@app.get("/preview", response_class=HTMLResponse)
async def preview(request: Request):
    return templates.TemplateResponse("preview.html", {"request": request})


# API エンドポイント
@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="XLSXファイルのみ対応しています")

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        data = extract_data_from_excel(ws)
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"ファイルの解析に失敗しました: {str(e)}")


@app.post("/api/preview")
async def generate_preview(data: SkillSheetData):
    return data





# Excelファイルからデータを抽出する関数
def extract_data_from_excel(worksheet):
    data = {
        "basic_info": {},
        "possible_tasks": {},
        "career_history": []
    }
    data["basic_info"]["name"] = worksheet.cell(row=2, column=4).value or ""
    data["basic_info"]["kana"] = worksheet.cell(row=3, column=4).value or ""
    gender_cell = worksheet.cell(row=2, column=6).value
    if gender_cell:
        if "男性" in str(gender_cell):
            data["basic_info"]["gender"] = "男性"
        elif "女性" in str(gender_cell):
            data["basic_info"]["gender"] = "女性"
        else:
            data["basic_info"]["gender"] = "その他"
    else:
        data["basic_info"]["gender"] = "回答しない"
    age_cell = worksheet.cell(row=3, column=6).value
    if age_cell and isinstance(age_cell, str) and "歳" in age_cell:
        try:
            data["basic_info"]["age"] = int(age_cell.replace("歳", "").strip())
        except (ValueError, TypeError):
            data["basic_info"]["age"] = None
    else:
        data["basic_info"]["age"] = None
    data["basic_info"]["nearest_station"] = worksheet.cell(row=4, column=4).value or ""
    exp_cell = worksheet.cell(row=5, column=4).value
    if exp_cell and isinstance(exp_cell, str) and "年" in exp_cell:
        try:
            data["basic_info"]["experience_years"] = int(exp_cell.replace("年", "").strip())
        except (ValueError, TypeError):
            data["basic_info"]["experience_years"] = None
    else:
        data["basic_info"]["experience_years"] = None
    return data


def calculate_career_duration(start_date_str: str, end_date_str: str) -> str:
    """業務期間を「X年Yヶ月」形式で計算する"""
    try:
        start_dt = datetime.strptime(start_date_str, "%Y-%m-%d")
    except (ValueError, TypeError):
        return ""

    if end_date_str == "current":
        end_dt = datetime.now()
    else:
        try:
            end_dt = datetime.strptime(end_date_str, "%Y-%m-%d")
        except (ValueError, TypeError):
            return ""

    if start_dt > end_dt:
        return ""

    total_months = (end_dt.year - start_dt.year) * 12 + (end_dt.month - start_dt.month) + 1
    if total_months <= 0:
        return ""

    years = total_months // 12
    months = total_months % 12

    duration_parts = []
    if years > 0:
        duration_parts.append(f"{years}年")
    if months > 0:
        duration_parts.append(f"{months}ヶ月")

    return "".join(duration_parts)
